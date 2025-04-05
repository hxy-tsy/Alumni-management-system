from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from django.db.models import Q
from .models import CouncilMeeting
from .serializers import CouncilMeetingSerializer
from users.models import User
from users.serializers import UserSerializer
import pandas as pd
from django.http import HttpResponse
import openpyxl
from datetime import datetime
import io
from alumni.models import AlumniProfile
import re

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def meeting_list(request):
    if request.method == 'GET':
        # 获取查询参数
        name = request.query_params.get('name', '')

        # 构建查询条件
        query = Q()
        if name:
            query &= Q(name__icontains=name)

        # 只显示当前用户创建的理事会
        if request.user.role == 'liaison':
            query &= Q(user=request.user)

        # 查询数据
        meetings = CouncilMeeting.objects.filter(query)

        # 分页
        paginator = StandardResultsSetPagination()
        paginated_meetings = paginator.paginate_queryset(meetings, request)
        serializer = CouncilMeetingSerializer(paginated_meetings, many=True)

        return paginator.get_paginated_response(serializer.data)

    elif request.method == 'POST':
        # 检查用户权限
        if request.user.role != 'liaison':
            return Response(
                {'error': '只有联络员可以创建理事会'},
                status=status.HTTP_403_FORBIDDEN
            )

        serializer = CouncilMeetingSerializer(
            data=request.data,
            context={'request': request}
        )
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def meeting_detail(request, pk):
    try:
        meeting = CouncilMeeting.objects.get(pk=pk)
        
        # 检查权限
        if request.user.role != 'liaison' or meeting.user != request.user:
            return Response(
                {'error': '没有权限执行此操作'},
                status=status.HTTP_403_FORBIDDEN
            )

        if request.method == 'PUT':
            # 如果已发送邀请，则不允许编辑
            if meeting.invitation_sent:
                return Response(
                    {'error': '已发送邀请的理事会无法编辑'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            serializer = CouncilMeetingSerializer(
                meeting,
                data=request.data,
                context={'request': request},
                partial=True
            )
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        elif request.method == 'DELETE':
            # 如果已发送邀请，则不允许删除
            if meeting.invitation_sent:
                return Response(
                    {'error': '已发送邀请的理事会无法删除'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            meeting.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

    except CouncilMeeting.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def meeting_batch_delete(request):
    # 检查用户权限
    if request.user.role != 'liaison':
        return Response(
            {'error': '只有联络员可以删除理事会'},
            status=status.HTTP_403_FORBIDDEN
        )

    ids = request.data.get('ids', [])
    if not ids:
        return Response(
            {"error": "未提供要删除的ID列表"},
            status=status.HTTP_400_BAD_REQUEST
        )

    # 只删除未发送邀请且属于当前用户的理事会
    meetings = CouncilMeeting.objects.filter(
        id__in=ids,
        user=request.user,
        invitation_sent=False
    )
    meetings.delete()

    return Response({"message": "批量删除成功"})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_invitation(request, pk):
    try:
        meeting = CouncilMeeting.objects.get(pk=pk)
        
        # 检查权限
        if request.user.role != 'liaison' or meeting.user != request.user:
            return Response(
                {'error': '没有权限执行此操作'},
                status=status.HTTP_403_FORBIDDEN
            )

        # 检查是否已发送邀请
        if meeting.invitation_sent:
            return Response(
                {'error': '邀请已发送'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # TODO: 在这里添加发送邀请的实际逻辑（如发送邮件或短信）

        # 更新邀请状态
        meeting.invitation_sent = True
        meeting.save()

        return Response({'message': '邀请发送成功'})

    except CouncilMeeting.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def withdraw_invitation(request, pk):
    try:
        meeting = CouncilMeeting.objects.get(pk=pk)
        
        # 检查权限
        if request.user.role != 'liaison' or meeting.user != request.user:
            return Response(
                {'error': '没有权限执行此操作'},
                status=status.HTTP_403_FORBIDDEN
            )

        # 检查是否已发送邀请
        if not meeting.invitation_sent:
            return Response(
                {'error': '邀请尚未发送'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # TODO: 在这里添加撤回邀请的实际逻辑（如发送取消通知）

        # 更新邀请状态
        meeting.invitation_sent = False
        meeting.save()

        return Response({'message': '邀请撤回成功'})

    except CouncilMeeting.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_meetings(request):
    try:
        # 检查用户权限
        if request.user.role != 'liaison':
            return Response(
                {'error': '只有联络员可以导入理事会'},
                status=status.HTTP_403_FORBIDDEN
            )

        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response(
                {"error": "请选择要导入的文件"},
                status=status.HTTP_400_BAD_REQUEST
            )

        df = pd.read_excel(excel_file)
        success_count = 0
        error_records = []

        for index, row in df.iterrows():
            try:
                # 基本会议信息
                meeting_data = {
                    'name': str(row['理事会名称']),
                    'content': str(row['内容']),
                    'location': str(row['地点']),
                    'meeting_time': row['召开时间'],
                    'invitees': []
                }

                # 处理邀请人员
                # Excel中的邀请人员应以逗号、分号或空格分隔（学号或姓名）
                if '邀请人员' in row and pd.notna(row['邀请人员']):
                    invitees_text = str(row['邀请人员'])
                    # 将文本分割成列表（支持逗号、分号、空格分隔）
                    invitee_ids = []
                    invitee_items = [item.strip() for item in re.split(r'[,;\s]+', invitees_text) if item.strip()]
                    
                    for item in invitee_items:
                        # 根据学号或姓名查找用户
                        query = Q()
                        # 如果是纯数字，假定为学号
                        if item.isdigit():
                            # 通过学号查找校友
                            alumni_profiles = AlumniProfile.objects.filter(student_id=item, is_graduated=1)
                            for profile in alumni_profiles:
                                invitee_ids.append(profile.user.id)
                        else:
                            # 通过姓名查找用户
                            users = User.objects.filter(username=item)
                            for user in users:
                                invitee_ids.append(user.id)
                    
                    meeting_data['invitees'] = invitee_ids

                serializer = CouncilMeetingSerializer(
                    data=meeting_data,
                    context={'request': request}
                )
                if serializer.is_valid():
                    serializer.save()
                    success_count += 1
                else:
                    error_records.append(f"第{index + 2}行: {serializer.errors}")

            except Exception as e:
                error_records.append(f"第{index + 2}行: {str(e)}")

        return Response({
            "success_count": success_count,
            "error_count": len(error_records),
            "errors": error_records
        })

    except Exception as e:
        return Response(
            {"error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_meetings(request):
    try:
        # 检查用户权限
        if request.user.role != 'liaison':
            return Response(
                {'error': '只有联络员可以导出理事会'},
                status=status.HTTP_403_FORBIDDEN
            )

        # 获取查询参数
        name = request.query_params.get('name', '')

        # 构建查询条件
        query = Q(user=request.user)  # 只导出当前用户的理事会
        if name:
            query &= Q(name__icontains=name)

        # 查询数据
        meetings = CouncilMeeting.objects.filter(query)

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '理事会列表'

        # 写入表头
        headers = ['理事会名称', '内容', '地点', '邀请人员', '召开时间', '是否已发送邀请']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 写入数据
        for row, meeting in enumerate(meetings, 2):
            ws.cell(row=row, column=1, value=meeting.name)
            ws.cell(row=row, column=2, value=meeting.content)
            ws.cell(row=row, column=3, value=meeting.location)
            # 修改获取邀请人员姓名的方式
            invitees = []
            for user in meeting.invitees.all():
                if hasattr(user, 'name') and user.name:
                    invitees.append(user.name)
                else:
                    invitees.append(user.username)
            ws.cell(row=row, column=4, value=', '.join(invitees))
            ws.cell(row=row, column=5, value=meeting.meeting_time.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=6, value='是' if meeting.invitation_sent else '否')

        # 保存到内存中
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=理事会列表_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return response

    except Exception as e:
        print(f"Export error: {str(e)}")  # 添加错误日志
        return Response(
            {"error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_graduated_alumni(request):
    """获取已毕业校友列表"""
    alumni = AlumniProfile.objects.filter(is_graduated=1).select_related('user')
    data = []
    for profile in alumni:
        data.append({
            'id': profile.user.id,  # 使用用户的ID而不是档案的ID
            'username': profile.user.username,
            'name': profile.user.name if hasattr(profile.user, 'name') else profile.user.username,
            'student_id': profile.student_id
        })
    return Response(data)