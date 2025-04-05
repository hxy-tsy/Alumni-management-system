from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes, parser_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from django.db.models import Count, Q, Avg, Sum
from .models import AlumniProfile, AlumniApplication
from .serializers import AlumniProfileSerializer, AlumniApplicationSerializer
from users.models import User
from rest_framework.parsers import MultiPartParser, FormParser
import os
import time
from django.conf import settings
import pandas as pd
from django.http import HttpResponse
import openpyxl
from io import BytesIO
import datetime
import random
from association.models import Association


class AlumniViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = AlumniProfile.objects.all()
    serializer_class = AlumniProfileSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        # 搜索过滤
        student_id = self.request.query_params.get('student_id', None)
        name = self.request.query_params.get('name', None)
        department = self.request.query_params.get('department', None)

        if student_id:
            queryset = queryset.filter(student_id__icontains=student_id)
        if name:
            queryset = queryset.filter(user__username__icontains=name)
        if department:
            queryset = queryset.filter(user__department__icontains=department)

        return queryset


class ApplicationViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = AlumniProfile.objects.filter(application_status=1)  # 只获取审核中的申请
    serializer_class = AlumniProfileSerializer

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        alumni = self.get_object()
        alumni.is_graduated = 1  # 设置为已毕业
        alumni.application_status = 2  # 设置为审批通过
        alumni.save()
        return Response({
            'status': 'approved',
            'message': '已批准该校友的毕业申请'
        })

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        alumni = self.get_object()
        alumni.is_graduated = 0  # 保持未毕业状态
        alumni.application_status = 3  # 设置为审批不通过
        alumni.save()
        return Response({
            'status': 'rejected',
            'message': '已拒绝该校友的毕业申请'
        })


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def alumni_list(request):
    if request.method == 'GET':
        # 获取查询参数
        student_id = request.query_params.get('student_id', '')
        name = request.query_params.get('name', '')
        # department = request.query_params.get('department', '')
        class_name = request.query_params.get('class_name', '')
        application_status = request.query_params.get('application_status', '')

        # 构建查询条件
        query = Q()  # 初始化一个空的Q对象
        department = request.user.department
        # 只有当参数有值时才添加到查询条件中
        if student_id:
            query &= Q(student_id__icontains=student_id)
        if name:
            query &= Q(user__username__icontains=name)
        if department:
            query &= Q(user__department__icontains=department)
        if class_name:
            query &= Q(class_name__icontains=class_name)
        if application_status:
            query &= Q(application_status=application_status)
        # 只显示和校友相同学院的校友信息
        # 查询数据
        # print("查询条件:", query)  # 调试信息
        alumni = AlumniProfile.objects.filter(query)
        # print("查询结果数量:", alumni.count())  # 调试信息
        serializer = AlumniProfileSerializer(alumni, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        print("收到的新增数据:", request.data)  # 打印接收到的数据
        serializer = AlumniProfileSerializer(data=request.data)
        if serializer.is_valid():
            # 直接使用序列化器的 create 方法创建
            alumni = serializer.save()
            return Response(AlumniProfileSerializer(alumni).data, status=status.HTTP_201_CREATED)
        print("验证错误:", serializer.errors)  # 打印验证错误信息
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def alumni_detail(request, pk):
    try:
        alumni = AlumniProfile.objects.get(pk=pk)
    except AlumniProfile.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = AlumniProfileSerializer(alumni)
        return Response(serializer.data)

    elif request.method == 'PUT':
        print("收到的请求数据:", request.data)  # 打印接收到的数据
        serializer = AlumniProfileSerializer(alumni, data=request.data, partial=True)
        print("序列化器验证前的数据:", serializer.initial_data)  # 打印初始数据
        if serializer.is_valid():
            # 更新用户信息
            user_data = request.data.get('user', {})
            if user_data:
                user = alumni.user
                for key, value in user_data.items():
                    if key != 'password':  # 不通过这里修改密码
                        setattr(user, key, value)
                user.save()

            # 更新校友信息
            alumni = serializer.save()
            return Response(AlumniProfileSerializer(alumni).data)
        print("序列化器验证错误:", serializer.errors)  # 打印验证错误信息
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        # 删除关联的用户
        alumni.user.delete()  # 这会级联删除校友信息
        return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def alumni_batch_delete(request):
    ids = request.data.get('ids', [])
    AlumniProfile.objects.filter(id__in=ids).delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def upload_avatar(request):
    if 'file' not in request.FILES:
        return Response({'error': '没有文件上传'}, status=status.HTTP_400_BAD_REQUEST)

    file = request.FILES['file']
    # 生成文件名
    filename = f"avatar_{request.user.id}_{int(time.time())}.jpg"
    # 文件保存路径
    filepath = os.path.join(settings.MEDIA_ROOT, 'avatars', filename)

    # 确保目录存在
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    # 保存文件
    with open(filepath, 'wb+') as destination:
        for chunk in file.chunks():
            destination.write(chunk)

    # 返回文件URL
    file_url = f"{settings.MEDIA_URL}avatars/{filename}"
    return Response({'url': file_url}, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def import_alumni(request):
    if 'file' not in request.FILES:
        return Response({'error': '请选择要导入的文件'}, status=status.HTTP_400_BAD_REQUEST)

    excel_file = request.FILES['file']
    try:
        # 读取Excel文件
        df = pd.read_excel(excel_file)
        success_count = 0
        error_count = 0
        errors = []

        for index, row in df.iterrows():
            try:
                # 创建用户
                user = User.objects.create(
                    username=row['姓名'],
                    phone=str(row['联系方式']),
                    department=row['学院'],
                    role='alumni',
                    password='123456'  # 默认密码
                )

                # 创建校友信息
                AlumniProfile.objects.create(
                    user=user,
                    student_id=str(row['学号']),
                    gender='male' if row['性别'] == '男' else 'female',
                    ethnicity=row['民族'],
                    birth_date=row['出生日期'],
                    address=row['通讯地址'],
                    class_name=row['班级'],
                    current_company=row['毕业去向'],
                    graduation_date=row['毕业日期']
                )
                success_count += 1
            except Exception as e:
                error_count += 1
                errors.append(f"第{index + 2}行导入失败：{str(e)}")

        return Response({
            'message': f'导入完成：成功{success_count}条，失败{error_count}条',
            'errors': errors
        })

    except Exception as e:
        return Response({'error': f'文件处理失败：{str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_alumni(request):
    try:
        # 获取所有校友数据
        alumni = AlumniProfile.objects.all().select_related('user')

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "校友信息"

        # 写入表头
        headers = ['学号', '姓名', '性别', '民族', '出生日期', '联系方式',
                   '学院', '班级', '毕业去向', '毕业日期', '通讯地址']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 写入数据
        for row, item in enumerate(alumni, 2):
            ws.cell(row=row, column=1, value=item.student_id)
            ws.cell(row=row, column=2, value=item.user.username)
            ws.cell(row=row, column=3, value='男' if item.gender == 'male' else '女')
            ws.cell(row=row, column=4, value=item.ethnicity)
            ws.cell(row=row, column=5, value=item.birth_date)
            ws.cell(row=row, column=6, value=item.user.phone)
            ws.cell(row=row, column=7, value=item.user.department)
            ws.cell(row=row, column=8, value=item.class_name)
            ws.cell(row=row, column=9, value=item.current_company)
            ws.cell(row=row, column=10, value=item.graduation_date)
            ws.cell(row=row, column=11, value=item.address)

        # 保存到内存中
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # 设置响应头
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=alumni_list_{datetime.date.today()}.xlsx'

        return response

    except Exception as e:
        return Response({'error': f'导出失败：{str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def apply_graduation(request, pk):
    try:
        alumni = AlumniProfile.objects.get(pk=pk)

        # 检查是否是本人的申请
        if request.user.id != alumni.user.id:
            return Response(
                {'error': '只能申请自己的毕业'},
                status=status.HTTP_403_FORBIDDEN
            )

        # 检查是否已经毕业
        if alumni.is_graduated == 1:
            return Response(
                {'error': '您已经毕业，无需再次申请'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 检查是否已经有待审核的申请
        if alumni.application_status == 1:
            return Response(
                {'error': '您已有一个待审核的毕业申请'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 更新毕业信息
        alumni.graduation_date = request.data.get('graduation_date')
        alumni.current_company = request.data.get('current_company')
        alumni.application_status = 1  # 设置为审核中状态
        alumni.save()

        return Response({
            'message': '毕业申请提交成功，请等待管理员审核',
            'graduation_date': alumni.graduation_date,
            'current_company': alumni.current_company,
            'status': 'pending'
        })

    except AlumniProfile.DoesNotExist:
        return Response(
            {'error': '未找到相关校友信息'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def current_alumni(request):
    """获取当前登录用户的校友信息"""
    try:
        alumni = AlumniProfile.objects.get(user=request.user)
        serializer = AlumniProfileSerializer(alumni)
        return Response(serializer.data)
    except AlumniProfile.DoesNotExist:
        return Response(
            {'error': '未找到校友信息'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def approve_application(request, pk):
    """批准校友毕业申请"""
    try:
        alumni = AlumniProfile.objects.get(pk=pk)

        # 检查是否是管理员
        if not request.user.is_staff:
            return Response(
                {'error': '只有管理员可以审批申请'},
                status=status.HTTP_403_FORBIDDEN
            )

        # 检查是否是待审核状态
        if alumni.application_status != 1:
            return Response(
                {'error': '该申请不在待审核状态'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 更新状态
        alumni.is_graduated = 1  # 设置为已毕业
        alumni.application_status = 2  # 设置为审批通过
        alumni.save()

        # 同时更新User表中的is_graduated字段
        user = alumni.user
        user.is_graduated = 1
        user.role = 'graduated_alumni'  # user设置为graduated_alumni表示已毕业的校友，alumni表示未毕业的校友
        user.save()

        return Response({
            'message': '已批准该校友的毕业申请',
            'status': 'approved'
        })

    except AlumniProfile.DoesNotExist:
        return Response(
            {'error': '未找到相关校友信息'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reject_application(request, pk):
    """拒绝校友毕业申请"""
    try:
        alumni = AlumniProfile.objects.get(pk=pk)

        # 检查是否是管理员
        if not request.user.is_staff:
            return Response(
                {'error': '只有管理员可以审批申请'},
                status=status.HTTP_403_FORBIDDEN
            )

        # 检查是否是待审核状态
        if alumni.application_status != 1:
            return Response(
                {'error': '该申请不在待审核状态'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 更新状态
        alumni.is_graduated = 0  # 保持未毕业状态
        alumni.application_status = 3  # 设置为审批不通过
        alumni.save()

        # 同时更新User表中的is_graduated字段
        user = alumni.user
        user.is_graduated = 0
        user.save()

        return Response({
            'message': '已拒绝该校友的毕业申请',
            'status': 'rejected'
        })

    except AlumniProfile.DoesNotExist:
        return Response(
            {'error': '未找到相关校友信息'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def statistics(request):
    """获取校友画像统计数据"""
    # 基础统计
    total_count = AlumniProfile.objects.count()
    gender_stats = AlumniProfile.objects.values('gender').annotate(count=Count('id'))
    male_count = next((item['count'] for item in gender_stats if item['gender'] == 'male'), 0)
    female_count = next((item['count'] for item in gender_stats if item['gender'] == 'female'), 0)

    # 院系分布
    department_stats = User.objects.filter(Q(role='alumni') | Q(role='graduated_alumni')).values('department').annotate(
        count=Count('id')
    ).order_by('-count')

    # 毕业年份分布
    graduation_year_stats = AlumniProfile.objects.filter(graduation_date__isnull=False).values('graduation_date__year').annotate(
        count=Count('id')
    ).order_by('graduation_date__year')

    # 就业去向分布（基于current_company字段分类）
    industry_stats = AlumniProfile.objects.exclude(current_company='').values('current_company').annotate(
        count=Count('id')
    ).order_by('-count')[:10]  # 只取前10个最多的公司类型

    # 分会分布
    association_stats = Association.objects.exclude(type='general').values('type').annotate(
        count=Count('id')
    ).order_by('-count')

    return Response({
        'total_count': total_count,
        'male_count': male_count,
        'female_count': female_count,
        'department_stats': list(department_stats),
        'graduation_year_stats': [{
            'year': item['graduation_date__year'],
            'count': item['count']
        } for item in graduation_year_stats],
        'industry_stats': [{
            'name': item['current_company'],
            'value': item['count']
        } for item in industry_stats],
        'association_stats': [{
            'name': dict(Association.TYPE_CHOICES).get(item['type']),
            'value': item['count']
        } for item in association_stats]
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def career_recommendation_data(request):
    """获取毕业去向推荐数据，包括MBTI相同校友和相同学院校友的毕业去向分布"""
    try:
        # 获取当前登录用户的校友信息
        try:
            alumni = AlumniProfile.objects.get(user=request.user)
        except AlumniProfile.DoesNotExist:
            return Response(
                {'error': '未找到您的校友信息'},
                status=status.HTTP_404_NOT_FOUND
            )

        # 获取请求中的毕业年份参数，默认为当前年份
        graduation_year = request.query_params.get('year', datetime.datetime.now().year)
        try:
            graduation_year = int(graduation_year)
        except (ValueError, TypeError):
            graduation_year = datetime.datetime.now().year

        # 获取MBTI相同的校友
        mbti_alumni = None
        mbti_stats = []
        if alumni.mbti:
            mbti_alumni = AlumniProfile.objects.filter(
                mbti=alumni.mbti,
                graduation_date__year=graduation_year,
                is_graduated=1
            ).exclude(id=alumni.id)
            
            # 直接计算毕业去向统计，不做分类
            if mbti_alumni.exists():
                mbti_stats = calculate_destination_stats(mbti_alumni)
        
        # 获取相同学院的校友
        same_department_alumni = AlumniProfile.objects.filter(
            user__department=request.user.department,
            graduation_date__year=graduation_year,
            is_graduated=1
        ).exclude(id=alumni.id)

        # 计算相同学院校友的毕业去向分布
        department_stats = []
        if same_department_alumni.exists():
            department_stats = calculate_destination_stats(same_department_alumni)

        return Response({
            'mbti_data': mbti_stats,
            'department_data': department_stats,
            'user_mbti': alumni.mbti,
            'user_department': request.user.department
        })

    except Exception as e:
        return Response(
            {'error': f'获取毕业去向数据失败: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# 辅助函数 - 直接计算去向统计，不做分类，不生成模拟数据
def calculate_destination_stats(queryset):
    """计算校友毕业去向统计数据，直接使用数据库中的原始去向"""
    # 排除空值并统计毕业去向分布
    destination_stats = queryset.exclude(current_company='').values('current_company').annotate(
        count=Count('id')
    ).order_by('-count')
    
    total = destination_stats.aggregate(Sum('count'))['count__sum'] or 0
    result = []
    
    # 将结果转换为前端需要的格式
    for item in destination_stats:
        company = item['current_company']
        count = item['count']
        value = round((count / total) * 100) if total > 0 else 0
        
        result.append({
            'name': company,
            'value': value
        })
    
    return result
