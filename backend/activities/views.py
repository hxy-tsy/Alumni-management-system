from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q, Count, Avg, F, ExpressionWrapper, fields
from django.db.models.functions import ExtractMonth
from rest_framework.pagination import PageNumberPagination
from django.shortcuts import get_object_or_404
from .models import Activity, ActivityMember, ActivityFeedback
from .serializers import ActivitySerializer, ActivityMemberSerializer, ActivityFeedbackSerializer
from users.models import User
from association.models import Association, AssociationMember
import pandas as pd
from django.http import HttpResponse
import openpyxl
from datetime import datetime
import io
import random



class StandardResultsSetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def activity_list(request):
    if request.method == 'GET':
        # 获取查询参数
        name = request.query_params.get('name', '')
        applicant_name = request.query_params.get('applicant_name', '')
        application_status = request.query_params.get('status', '')

        # 构建查询条件
        query = Q()

        if name:
            query &= Q(name__icontains=name)
        if applicant_name:
            query &= Q(applicant_name__icontains=applicant_name)
        if application_status:
            query &= Q(status=application_status)

        # 查询数据
        activities = Activity.objects.filter(query).order_by('-apply_time')

        # 使用分页
        paginator = StandardResultsSetPagination()
        paginated_activities = paginator.paginate_queryset(activities, request)

        # 序列化多个对象
        serializer = ActivitySerializer(paginated_activities, many=True)

        # 返回分页响应
        return paginator.get_paginated_response(serializer.data)

    elif request.method == 'POST':
        print("收到的新增数据:", request.data)
        
        # 1. 验证申请人是否是校友联络员
        applicant_name = request.data.get('applicant_name')
        organization = request.data.get('organization')
        
        try:
            # 查找申请人用户
            applicant = User.objects.get(username=applicant_name)
            
            # 检查是否是校友联络员
            if applicant.role != 'liaison':
                return Response(
                    {"error": "申请人必须是校友联络员"},
                    status=status.HTTP_400_BAD_REQUEST
                )
                
            # 2. 验证举办组织（校友会）是否存在且申请人是否是该校友会的联络员
            association = Association.objects.filter(name=organization).first()
            if not association:
                return Response(
                    {"error": "举办组织不存在"},
                    status=status.HTTP_400_BAD_REQUEST
                )
                
            # 检查申请人是否是该校友会的联络员
            is_liaison = Association.objects.filter(name=organization, leader_id=applicant.id).exists()
            
            if not is_liaison:
                return Response(
                    {"error": "申请人不是该校友会的联络员"},
                    status=status.HTTP_400_BAD_REQUEST
                )
                
            # 验证通过，创建活动
            serializer = ActivitySerializer(data=request.data)
            if serializer.is_valid():
                activity = serializer.save(
                    status=1,  # 设置状态为"申请中"
                )
                return Response(ActivitySerializer(activity).data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
        except User.DoesNotExist:
            return Response(
                {"error": "申请人不存在"},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


@api_view(['PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def activity_detail(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    applicant_name = request.data.get('applicant_name')
    organization = request.data.get('organization')
    
    # 检查是否为活动创建者
    if activity.status == 2:  # 如果活动已通过审核
        return Response({"error": "已通过审核的活动不能修改或删除"}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'PUT':

        try:
            # 查找申请人用户
            applicant = User.objects.get(username=applicant_name)
            # 检查是否是校友联络员
            if applicant.role != 'liaison':
                return Response(
                    {"error": "申请人必须是校友联络员"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 2. 验证举办组织（校友会）是否存在且申请人是否是该校友会的联络员
            association = Association.objects.filter(name=organization).first()
            if not association:
                return Response(
                    {"error": "举办组织不存在"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 检查申请人是否是该校友会的联络员
            is_liaison = Association.objects.filter(name=organization, leader_id=applicant.id).exists()

            if not is_liaison:
                return Response(
                    {"error": "申请人不是该校友会的联络员"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            serializer = ActivitySerializer(activity, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except User.DoesNotExist:
            return Response(
                {"error": "申请人不存在"},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    elif request.method == 'DELETE':
        activity.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def activity_batch_delete(request):
    try:
        activity_ids = request.data.get('ids', [])
        if not activity_ids:
            return Response({"error": "未提供要删除的活动ID"}, status=status.HTTP_400_BAD_REQUEST)

        # 获取所有要删除的活动
        activities = Activity.objects.filter(id__in=activity_ids)
        
        # 检查是否有已通过审核的活动
        approved_activities = activities.filter(status=2)
        if approved_activities.exists():
            return Response({"error": "选中的活动中包含已通过审核的活动，无法删除"}, 
                          status=status.HTTP_403_FORBIDDEN)

        # 执行删除
        activities.delete()
        return Response({"message": "批量删除成功"}, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_activities(request):
    try:
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"error": "请选择要导入的文件"}, status=status.HTTP_400_BAD_REQUEST)

        # 读取Excel文件
        df = pd.read_excel(excel_file)
        
        success_count = 0
        error_records = []

        for index, row in df.iterrows():
            try:
                # 验证申请人
                applicant_name = str(row['申请人姓名'])
                applicant = User.objects.get(username=applicant_name)
                
                # 验证校友会
                organization = str(row['举办组织'])
                association = Association.objects.filter(name=organization).first()
                
                if not association:
                    error_records.append(f"第{index + 2}行: 举办组织不存在")
                    continue
                    
                # 验证申请人是否是该校友会的联络员
                is_liaison = Association.objects.filter(name=organization, leader_id=applicant.id).exists()
                if not is_liaison:
                    error_records.append(f"第{index + 2}行: 申请人不是该校友会的联络员")
                    continue

                # 创建活动
                activity = Activity(
                    name=str(row['活动名称']),
                    description=str(row['活动介绍']),
                    applicant_name=applicant_name,
                    phone=str(row['联系电话']),
                    organization=organization,
                    venue=str(row['场地设施']),
                    event_time=row['活动举办时间'],
                    status=1  # 设置为申请中
                )
                activity.save()
                success_count += 1
                
            except User.DoesNotExist:
                error_records.append(f"第{index + 2}行: 申请人不存在")
            except Exception as e:
                error_records.append(f"第{index + 2}行: {str(e)}")

        result = {
            "success_count": success_count,
            "error_count": len(error_records),
            "errors": error_records
        }
        
        return Response(result, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_activities(request):
    try:
        # 获取查询参数
        name = request.query_params.get('name', '')
        applicant_name = request.query_params.get('applicant_name', '')
        application_status = request.query_params.get('status', '')

        # 构建查询条件
        query = Q()
        if name:
            query &= Q(name__icontains=name)
        if applicant_name:
            query &= Q(applicant_name__icontains=applicant_name)
        if application_status:
            query &= Q(status=application_status)

        # 查询数据
        activities = Activity.objects.filter(query).order_by('-apply_time')

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '校友活动列表'

        # 写入表头
        headers = ['活动名称', '活动介绍', '申请人姓名', '联系电话', '举办组织', 
                  '场地设施', '申请时间', '活动举办时间', '审批状态']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 写入数据
        for row, activity in enumerate(activities, 2):
            ws.cell(row=row, column=1, value=activity.name)
            ws.cell(row=row, column=2, value=activity.description)
            ws.cell(row=row, column=3, value=activity.applicant_name)
            ws.cell(row=row, column=4, value=activity.phone)
            ws.cell(row=row, column=5, value=activity.organization)
            ws.cell(row=row, column=6, value=activity.venue)
            ws.cell(row=row, column=7, value=activity.apply_time.strftime('%Y-%m-%d %H:%M:%S') if activity.apply_time else '')
            ws.cell(row=row, column=8, value=activity.event_time.strftime('%Y-%m-%d %H:%M:%S') if activity.event_time else '')
            ws.cell(row=row, column=9, value=activity.get_status_display())

        # 保存到内存中
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # 设置响应头
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=activities_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def approve_activity(request, pk):
    # 检查用户是否是管理员
    if not request.user.is_staff:
        return Response({"error": "只有管理员可以审批活动"}, status=status.HTTP_403_FORBIDDEN)
        
    activity = get_object_or_404(Activity, pk=pk)
    
    # 检查活动是否处于申请中状态
    if activity.status != 1:
        return Response({"error": "只能审批处于申请中状态的活动"}, status=status.HTTP_400_BAD_REQUEST)
        
    # 更新活动状态为已通过
    activity.status = 2
    activity.save()
    
    return Response({"message": "活动审批通过"}, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reject_activity(request, pk):
    # 检查用户是否是管理员
    if not request.user.is_staff:
        return Response({"error": "只有管理员可以审批活动"}, status=status.HTTP_403_FORBIDDEN)
        
    activity = get_object_or_404(Activity, pk=pk)
    
    # 检查活动是否处于申请中状态
    if activity.status != 1:
        return Response({"error": "只能审批处于申请中状态的活动"}, status=status.HTTP_400_BAD_REQUEST)
        
    # 更新活动状态为已拒绝
    activity.status = 3
    activity.save()
    
    return Response({"message": "已拒绝该活动申请"}, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reapply_activity(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    
    # 检查是否为活动创建者
    if activity.applicant_name != request.user.username:
        return Response({"error": "只有活动申请人可以重新申请"}, status=status.HTTP_403_FORBIDDEN)
    
    # 检查活动是否被拒绝
    if activity.status != 3:
        return Response({"error": "只有被拒绝的活动可以重新申请"}, status=status.HTTP_400_BAD_REQUEST)
    
    # 更新活动状态为申请中
    activity.status = 1
    activity.save()
    
    return Response({"message": "重新申请成功"}, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def activity_registration_list(request):
    # 获取查询参数
    name = request.query_params.get('name', '')
    user = request.user
    # 构建查询条件
    query = Q(status=2)  # 只显示已通过审核的活动
    if name:
        query &= Q(name__icontains=name)

    # 查询数据
    activities = Activity.objects.filter(query).order_by('-apply_time')

    activityMemberQuery = Q(user_id=user.id)
    ActivityMember.objects.filter(activityMemberQuery)
    # 使用分页
    paginator = StandardResultsSetPagination()
    paginated_activities = paginator.paginate_queryset(activities, request)

    # 序列化多个对象
    serializer = ActivitySerializer(paginated_activities, many=True, context={'request': request})

    # 返回分页响应
    return paginator.get_paginated_response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def register_activity(request, pk):
    activity = get_object_or_404(Activity, pk=pk)
    
    # 检查活动是否已通过审核
    if activity.status != 2:
        return Response({"error": "只能报名已通过审核的活动"}, status=status.HTTP_400_BAD_REQUEST)
    
    # 检查是否已经报名
    if ActivityMember.objects.filter(activity=activity, user=request.user).exists():
        return Response({"error": "您已经报名过该活动"}, status=status.HTTP_400_BAD_REQUEST)
    
    # 创建报名记录
    registration = ActivityMember(
        activity=activity,
        user=request.user,
        status=1  # 设置为申请中状态
    )
    registration.save()
    
    return Response({"message": "报名成功"}, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_activities(request):
    """获取当前用户报名的活动列表"""
    queryset = Activity.objects.filter(members__user=request.user)
    
    # 搜索过滤
    name = request.query_params.get('name', '')
    if name:
        queryset = queryset.filter(name__icontains=name)
    
    # 分页
    paginator = StandardResultsSetPagination()
    page = paginator.paginate_queryset(queryset, request)
    
    serializer = ActivitySerializer(page, many=True, context={'request': request})
    return paginator.get_paginated_response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def cancel_registration(request, pk):
    """取消活动报名"""
    try:
        activity_member = ActivityMember.objects.get(activity_id=pk, user=request.user)
        activity_member.delete()
        return Response({'message': '取消报名成功'})
    except ActivityMember.DoesNotExist:
        return Response({'error': '未找到报名记录'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=400)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def activity_planning_stats(request):
    # 获取所有已批准的活动
    approved_activities = Activity.objects.filter(
        status=2,
        event_time__isnull=False
    )

    # 计算基础统计数据
    total_activities = approved_activities.count()
    total_participants = ActivityMember.objects.filter(activity__in=approved_activities).count()
    avg_participants = round(total_participants / total_activities if total_activities > 0 else 0)

    # 获取每个活动的名称、参与人数和平均评分
    activity_stats = []
    for activity in approved_activities:
        participant_count = ActivityMember.objects.filter(activity=activity).count()
        # 获取活动的平均评分
        avg_rating = ActivityFeedback.objects.filter(activity=activity).aggregate(Avg('rating'))['rating__avg']
        # 如果没有评分，设置默认值为0
        rating = round(avg_rating, 1) if avg_rating is not None else 0
        
        activity_stats.append({
            'name': activity.name,
            'participants': participant_count,
            'rating': rating
        })

    # 按月份统计活动数量
    monthly_stats = [0] * 12  # 初始化12个月的数组
    for activity in approved_activities:
        month = activity.event_time.month
        if month:
            monthly_stats[month - 1] += 1

    best_activity_name = ""
    best_activity_rating = -1
    for activity in activity_stats:
        if activity['rating'] > best_activity_rating:
            best_activity_rating = activity['rating']
            best_activity_name = activity['name']
    # 生成推荐建议
    recommendations = [
        f"最适合举办活动的月份是{monthly_stats.index(max(monthly_stats)) + 1}月，历史活动数量最多。",
        f"平均每个活动有{avg_participants}人参与，建议根据场地容量合理安排人数。",
        f"{best_activity_name}活动的好评度最高，建议多举办类似活动。"
    ]

    return Response({
        'total_activities': total_activities,
        'avg_participants': avg_participants,
        'activity_stats': activity_stats,
        'monthly_stats': monthly_stats,
        'recommendations': recommendations
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def submit_feedback(request, pk):
    """提交活动反馈"""
    try:
        activity = get_object_or_404(Activity, pk=pk)
        
        # 检查用户是否参加了该活动
        if not ActivityMember.objects.filter(activity=activity, user=request.user).exists():
            return Response({'error': '您没有参加此活动，无法提交反馈'}, status=status.HTTP_403_FORBIDDEN)
            
        # 获取或创建反馈记录
        feedback, created = ActivityFeedback.objects.get_or_create(
            activity=activity,
            user=request.user,
            defaults={
                'rating': request.data.get('rating', 5),
                'comment': request.data.get('comment', '')
            }
        )
        
        if not created:
            # 更新现有反馈
            feedback.rating = request.data.get('rating', feedback.rating)
            feedback.comment = request.data.get('comment', feedback.comment)
            feedback.save()
            
        serializer = ActivityFeedbackSerializer(feedback)
        return Response(serializer.data, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_feedback(request, pk):
    """获取活动反馈"""
    try:
        activity = get_object_or_404(Activity, pk=pk)
        feedback = ActivityFeedback.objects.filter(activity=activity, user=request.user).first()
        
        if feedback:
            serializer = ActivityFeedbackSerializer(feedback)
            return Response(serializer.data)
        else:
            # 返回空数据而不是404
            return Response({
                'activity': pk,
                'rating': 5,
                'comment': '',
                'exists': False
            })
            
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
