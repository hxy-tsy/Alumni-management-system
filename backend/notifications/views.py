# backend/notifications/views.py

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.pagination import PageNumberPagination
from django.db.models import Q
from .models import Notification
from .serializers import NotificationSerializer
from users.models import User
from users.serializers import UserSerializer
import pandas as pd
from django.http import HttpResponse
import openpyxl
from datetime import datetime
import io
from alumni.models import AlumniProfile
from alumni.serializers import AlumniProfileSerializer


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def greeting_list(request):
    if request.method == 'GET':
        # 获取查询参数
        content = request.query_params.get('content', '')
        type_param = request.query_params.get('type', '')
        status_param = request.query_params.get('status', '')

        # 构建查询条件
        query = Q()
        if content:
            query &= Q(content__icontains=content)
        if type_param:
            query &= Q(type=type_param)
        if status_param and status_param.isdigit():
            query &= Q(status=int(status_param))

        # 根据用户角色过滤数据
        if request.user.role == 'graduated_alumni':  # 如果不是管理员
            query &= Q(receivers=request.user)  # 只显示接收者包含当前用户的通知，且发送成功的
            query &= Q(status=0)

        # 查询数据
        greetings = Notification.objects.filter(query).order_by('-send_time')

        # 分页
        paginator = StandardResultsSetPagination()
        paginated_greetings = paginator.paginate_queryset(greetings, request)
        serializer = NotificationSerializer(paginated_greetings, many=True)

        # 在序列化数据中添加type_display字段
        response = paginator.get_paginated_response(serializer.data)
        return response

    elif request.method == 'POST':
        serializer = NotificationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(
                sender=request.user,
                send_time=datetime.now(),  # 自动设置发送时间为当前时间
                status=0  # 设置初始状态
            )
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def greeting_detail(request, pk):
    try:
        greeting = Notification.objects.get(pk=pk)
    except Notification.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    if request.method == 'PUT':
        original_send_time = greeting.send_time
        serializer = NotificationSerializer(greeting, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(send_time=original_send_time)
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        greeting.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def greeting_batch_delete(request):
    ids = request.data.get('ids', [])
    if not ids:
        return Response({"error": "未提供要删除的ID列表"}, status=status.HTTP_400_BAD_REQUEST)

    Notification.objects.filter(id__in=ids).delete()
    return Response({"message": "批量删除成功"})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_greetings(request):
    try:
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"error": "请选择要导入的文件"}, status=status.HTTP_400_BAD_REQUEST)

        df = pd.read_excel(excel_file)
        success_count = 0
        error_records = []
        type_map = {
            '新闻通知': 'news',
            '活动通知': 'activity',
            '节日祝福': 'greeting'
        }
        for index, row in df.iterrows():
            try:
                # 处理接收者字段
                receivers_str = str(row.get('接收者', ''))
                receiver_usernames = [username.strip() for username in
                                      receivers_str.split(',')] if receivers_str else []

                # 查找接收者用户
                receivers = User.objects.filter(username__in=receiver_usernames)
                if not receivers.exists() and receiver_usernames:
                    error_records.append(f"第{index + 2}行: 未找到指定的接收者用户")
                    continue

                type_pram = type_map[row['短信类型']]
                notification = Notification(
                    title=str(row['标题']),
                    content=str(row['内容']),
                    type=type_pram,
                    sender=request.user,
                    status=0,
                    send_time=datetime.now()  # 设置发送时间
                )
                notification.save()

                # 设置接收者
                if receivers.exists():
                    notification.receivers.set(receivers)

                success_count += 1
            except Exception as e:
                error_records.append(f"第{index + 2}行: {str(e)}")

        return Response({
            "success_count": success_count,
            "error_count": len(error_records),
            "errors": error_records
        })

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_greetings(request):
    try:
        # 获取查询参数
        content = request.query_params.get('content', '')
        status_param = request.query_params.get('status', '')
        type_param = request.query_params.get('type', '')
        # 构建查询条件
        query = Q()
        if content:
            query &= Q(content__icontains=content)
        if status_param:
            query &= Q(status=status_param)
        if type_param:
            query &= Q(type=type_param)

        # 查询数据
        greetings = Notification.objects.filter(query).order_by('-send_time')

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '节日祝福'

        # 写入表头
        headers = ['标题', '短信类型', '内容', '发送状态', '发送时间', '接收者']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 写入数据
        for row, greeting in enumerate(greetings, 2):
            ws.cell(row=row, column=1, value=greeting.title)
            ws.cell(row=row, column=2, value=greeting.get_type_display())
            ws.cell(row=row, column=3, value=greeting.content)
            ws.cell(row=row, column=4, value=greeting.get_status_display())
            ws.cell(row=row, column=5,
                    value=greeting.send_time.strftime('%Y-%m-%d %H:%M:%S') if greeting.send_time else '')
            # 获取并拼接所有接收者的用户名
            receivers = greeting.receivers.all()
            receivers_str = ', '.join([user.username for user in receivers])
            ws.cell(row=row, column=6, value=receivers_str)

        # 保存到内存中
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response[
            'Content-Disposition'] = f'attachment; filename=greetings_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return response

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def graduated_alumni_list(request):
    """获取已毕业校友列表"""
    alumni = AlumniProfile.objects.filter(is_graduated=1).select_related('user')
    data = []
    for profile in alumni:
        data.append({
            'id': profile.user.id,  # 使用用户的ID而不是档案的ID
            'username': profile.user.username,
            'name': profile.user.name if hasattr(profile.user, 'name') else profile.user.username,
            'student_id': profile.student_id
        })
    return Response(data)
